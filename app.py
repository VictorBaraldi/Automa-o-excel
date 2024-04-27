import openpyxl
import time
import os
from dotenv import load_dotenv
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

book = openpyxl.load_workbook('dados.xlsx')
dados = book['Planilha1']
lista_de_busca = []

def init():
  for rows in dados.iter_rows(min_row=23, max_row=26):
      if isinstance(rows[15].value, str):
        if rows[15].value.lower() == 'sim': lista_de_busca.append(rows[3].value.replace('-', ''))
      else:
        continue
  try:
    planilha_excel = book['Dados']
  except:
    book.create_sheet('Dados')
    planilha_excel = book['Dados']
  finally:  
    planilha_excel.append(['MÊS'] + ['GRUPO'] + ['COTA'] + ['CPF'] + ['NOME'] + ['DATA NASCIMENTO'] + ['NOME DA MÃE'] + ['CRÉDITO'] + ['Nº PARC'] + ['VALOR PARC'] + ['Nº PARC'] + ['VALOR PARC'] + ['PAGO ANOS ANTERIORES'] + ['TOTAL PAGO'] + ['TELEFONES'])
    login()
    return planilha_excel

chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
servico = Service(ChromeDriverManager().install())
navegador = webdriver.Chrome(service=servico, options=chrome_options)

def login():
  load_dotenv()
  email_login = os.getenv("EMAIL_LOGIN")
  senha_login = os.getenv("SENHA_LOGIN")
  navegador.get('https://lemitti.com/auth/login')
  navegador.find_element('xpath', '//*[@id="email"]').send_keys(f'{email_login}')
  navegador.find_element('xpath', '//*[@id="password"]').send_keys(f'{senha_login}')
  button = navegador.find_element('xpath', '//*[@id="content"]/div/div/div/form/div/div/div[4]/button')
  navegador.execute_script("arguments[0].click();", button)

planilha_excel = init()    

def tratar_dados(dado):
  if len(dado) == 14:
    cnpj = '{}.{}.{}/{}.{}-{}'.format(dado[:3], dado[3:6], dado[6:8], dado[8:11], dado[11:13], dado[13:])
    return cnpj
  elif len(dado) == 11:
    cpf = '{}.{}.{}-{}'.format(dado[:3], dado[3:6], dado[6:9], dado[9:])
    return cpf
  else:
    return 1111111111

def busca_dados(dado):
  navegador.get('https://lemitti.com/queries')
  navegador.find_element('xpath', '//*[@id="document"]').send_keys(dado)
  button = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[5]/div/div/div/form/div/div/span/button/i')
  navegador.execute_script("arguments[0].click();", button)


def coleta_dados(dado):
  telefones_raiz = busca_telefone() 

  if len(dado) == 14:
    data = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div/table/tbody/tr[5]/td').text
    nome = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div/table/tbody/tr[2]/td').text
    informacao_socios = 'nenhum sócio encontrado'
    quantidade_socios = verifica_quantidade('//*[@id="content"]/div[3]/div[4]/div/div[5]/div[2]/div/table/tbody/tr/td/em', '//*[@id="content"]/div[3]/div[4]/div/div[5]/div[2]/div/table/tbody/tr', 'nenhum sócio encontrado')
    print(quantidade_socios)
    if quantidade_socios > 0:
      informacao_socios = ''
      for i in socios_cnpj(quantidade_socios):
        busca_dados(i)
        telefones = busca_telefone()
        if len(i) == 14:
          nome_socio = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div[1]/table/tbody/tr[1]/td').text
        if len(i) == 19:
          nome_socio = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div/table/tbody/tr[2]/td').text
        informacao_socios += f'{nome_socio}: {telefones}   '
        time.sleep(12)
    planilha_excel.append([None]*4 + [nome] + [data] + [informacao_socios] + [None]*7 + [telefones_raiz] + ['não'])  

  if (len(dado)) == 11:
    data = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div[1]/table/tbody/tr[4]/td').text
    nome = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div[1]/table/tbody/tr[1]/td').text
    mae = navegador.find_element('xpath', '//*[@id="content"]/div[3]/div[4]/div/div[3]/div[2]/div[1]/table/tbody/tr[6]/td').text
    planilha_excel.append([None]*4 + [nome] + [data] + [mae] + [None]*7 + [telefones_raiz] + ['não'])     
    
  book.save('dados.xlsx')

def verifica_quantidade(caminho_texto, caminho_telefone, texto_comparacao):
  try:
    if navegador.find_element('xpath', f'{caminho_texto}').text == f'{texto_comparacao}':
      quantidade = 0 
      return quantidade
  except: 
    elementos = navegador.find_elements('xpath', f'{caminho_telefone}')
    quantidade_de_elementos = len(elementos)
    return quantidade_de_elementos
  
def busca_telefone():
  telefones = ''                                          
  quantidade_telefones_moveis = verifica_quantidade('//*[@id="content"]/div[3]/div[4]/div/div[6]/div[2]/div/table/tbody/tr/td/em', '//*[@id="content"]/div[3]/div[4]/div/div[6]/div[2]/div/table/tbody/tr', 'nenhum telefone encontrado')
  quantidade_telefones_fixos = verifica_quantidade('//*[@id="content"]/div[3]/div[4]/div/div[7]/div[2]/div/table/tbody/tr/td/em', '//*[@id="content"]/div[3]/div[4]/div/div[7]/div[2]/div/table/tbody/tr','nenhum telefone encontrado')
  if quantidade_telefones_moveis == 0 and quantidade_telefones_fixos == 0: telefones = 'nenhum telefone encontrado'
  if quantidade_telefones_moveis > 0: 
    for i in range(quantidade_telefones_moveis):
      telefone = navegador.find_element('xpath', f'//*[@id="content"]/div[3]/div[4]/div/div[6]/div[2]/div/table/tbody/tr[{i + 1}]/td[1]/a[1]').text
      telefones = telefones + '  ' + telefone
  if quantidade_telefones_fixos > 0:    
    for i in range(quantidade_telefones_fixos):
      telefone = navegador.find_element('xpath', f'//*[@id="content"]/div[3]/div[4]/div/div[7]/div[2]/div/table/tbody/tr[{i + 1}]/td[1]/a[1]').text
      telefones = telefones + '  ' + telefone
  return telefones

def socios_cnpj(quantidade_socios):
  dados_socios = []
  for i in range(quantidade_socios):
    informacao_do_socio = navegador.find_element('xpath', f'//*[@id="content"]/div[3]/div[4]/div/div[5]/div[2]/div/table/tbody/tr[{i + 1}]/td[2]/a').text
    dado_limpo = informacao_do_socio.replace('.', '').replace('-', '').replace('/', '')
    dados_formatado = tratar_dados(dado_limpo)
    dados_socios.append(dados_formatado)
  print(len(dados_socios))
  return dados_socios



for linha in lista_de_busca:  
  informacao = tratar_dados(linha)
  busca_dados(informacao)
  time.sleep(12)
  try:
    coleta_dados(linha)
  except:
    planilha_excel.append([None]*4 + ['Erro no processamento'] + [None]*10 + ['sim'])
    continue  


